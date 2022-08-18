import time
from turtle import title
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.common.exceptions import NoSuchElementException
from selenium.common.exceptions import StaleElementReferenceException
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions
from openpyxl import Workbook,load_workbook
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.chrome.options import Options
import os
import json

def load_category (filename):
    category_wb = load_workbook(filename, read_only=True)
    category_sw = category_wb.active
    category_info = [[cell.value for cell in sub_category ] for sub_category in category_sw.rows]
    category_info.append(category_info)
    return category_info

def get_info(category , browser):
    browser.get(category)
    time.sleep(10)
    categories = []
    card = browser.find_elements(By.CLASS_NAME, 'pub-visual-navigation__item is-small')
    print(len(card))
    for cards in card:
        title = browser.find_element(By.XPATH,'/html/body/div[1]/div/div/div[2]/div[1]/div[2]/div[2]/div/span/a').text
        link = card.get_attribute('href')
        print(title,link)
        categories = [title, link]
        categories.append(categories)
    return categories

def save_info(info, file_name):
    wb = Workbook()
    ws = wb.active
    for row in info:
        ws.append(row)
    wb.save(file_name) 

if __name__ == '__main__':
    path_driver = 'chromedriver.exe'
    browser = webdriver.Chrome(executable_path = path_driver)
    link_info = []
    print('starting')
    info = load_category('large_category.xlsx')
    i=1;
    for category in info:
        
        print(f"{i} : {category[0]}")
        link_info = get_info(category[0], browser)
        i+=1
    save_info(link_info, 'links.xlsx')
    