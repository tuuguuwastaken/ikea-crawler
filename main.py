import time
from pathlib import Path
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.common.exceptions import NoSuchElementException
from selenium.common.exceptions import StaleElementReferenceException
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions
from openpyxl import Workbook,load_workbook
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.chrome.options import Options
import os
import json

def load_category (filename):
    category_wb = load_workbook(filename, read_only=True)
    category_sw = category_wb.active
    category_info = [[cell.value for cell in sub_category ] for sub_category in category_sw.rows]
    category_info.append(category_info)
    return category_info
def scroll_bot():
    SCROLL_PAUSE_TIME = 2.0
    last_height = driver.execute_script("return document.body.scrollHeight")
    print(last_height)

    while True:
        time.sleep(SCROLL_PAUSE_TIME)

        driver.execute_script("window.scrollTo(0, document.body.scrollHeight);")
        time.sleep(SCROLL_PAUSE_TIME)
        new_height = driver.execute_script("return document.body.scrollHeight")
        print(new_height)
        if new_height == last_height:
            break
        last_height = new_height

def get_info_from_url(browser, sub_category):
    
    print(sub_category)
    products_info = []
    
    while(True):
            print('while loop')
            try:
                element = WebDriverWait(driver,80).until(
                    EC.presence_of_element_located((By.CLASS_NAME,'per-product-link-wrapper') and (By.CLASS_NAME,'per-product-link-wrapper'))
                )
                
            except:
                browser.refresh()
                pass
            scroll_bot()       
            cards = browser.find_elements(By.CLASS_NAME,'per-product-link-wrapper')
            print('yeet')
            for card in cards :
                        try:
                            element = WebDriverWait(driver,15).until(
                                EC.presence_of_element_located((By.XPATH,'/html/body/div[1]/div/div/div[2]/div[1]/div[2]/div[2]/div/span[2]/a') or (By.CLASS_NAME, 'ivu-breadcrumb-item-link'))
                            )
                        except:
                            browser.refresh()
                            continue
                        else:
                            pass  
                        if(len(browser.find_elements(By.CLASS_NAME,'per-product-link-wrapper')) < 0 ):
                            print("nothing skipping")
                            cards += 1
                            continue
                        else :
                            WebDriverWait(driver, 40)
                            title = browser.find_element(By.CLASS_NAME, 'ivu-breadcrumb-item-link').text or browser.find_element(By.XPATH,'/html/body/div[1]/div/div/div[2]/div[1]/div[2]/div[3]/div[1]/div/h1/div')
                            print(title)
                            product_name = card.find_element(By.CLASS_NAME, "withprice-title").text
                            product_desc = card.find_element(By.CLASS_NAME, 'withprice-commit').text
                            product_price = card.find_element(By.CLASS_NAME,'product-price-wrapper').text 
                            
                            products_images = card.find_element(By.TAG_NAME, 'img').get_attribute("src")
                            print(product_name, product_desc ,product_price,products_images)
                            product_url = card.get_attribute('href')
                            product_code = product_url.split("-")[-1].strip("s/")
                            prod = {

                                'sub_category': title,
                                'product_info': [{
                                'name': product_name,
                                    'desc': product_desc,
                                    'price': product_price,
                                    'product url': product_url,
                                    'product code': product_code,
                                    'product_image': products_images
                                }]
                            }
                            products_info.append(prod)  
            with open(f'JSON/{sub_category}/{title}.json', 'w') as outfile:
                json.dump(products_info, outfile)
            break

def save_info(info, file_name):
    wb = Workbook()
    ws = wb.active
    for row in info:
        ws.append(row)
    wb.save(file_name) 

if __name__ == "__main__":
    path_driver = 'chromedriver.exe'
    dir_path = os.getcwd()
    chrome_options = Options()
    chrome_options.add_argument(f'user-data-dir={dir_path}/selenium')
    driver = webdriver.Chrome(executable_path = path_driver, chrome_options=chrome_options)
    info = load_category('ikea-link copy 2.xlsx')
    print(info)
    product_info = []
    cwd = os.getcwd()
    JSON_PATH = os.path.join(cwd,'JSON')
    i = 1
    
    for sub_category in info:
        Path_for_json = os.path.join(JSON_PATH,str(sub_category[0]))
        try:
            os.mkdir(Path_for_json)
        except:
            
            pass
    for sub_category in info:
        print(f" checking and submitting {sub_category[0]} ({i}/{len(info)})")
        link = sub_category[1]
        driver.get(link)
        product_info = get_info_from_url(driver, sub_category[0])
        print(sub_category[0]+ ' ' + sub_category[1])
        i+=1
        # product_info.append(product_info)
        
        
        
    # save_info(product_info, "catergory.xlsx")
    # with open(f'json.json', 'w') as outfile:
    #      json.dump(product_info, outfile)
    
    # print(sub_category)
    driver.close()





